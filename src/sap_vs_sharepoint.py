import pandas as pd
import numpy as np
import pyttsx3
import os
import shutil
import time

from datetime import date
from openpyxl import load_workbook

start = time.time()

print()

today = date.today().strftime("%d%m%y")
folder_exp = f'C:/Users/J1049122/Desktop/Station Data/Master-Data/export/SAP_vs_SHAREPOINT/sortie_{today}'

if os.path.exists(folder_exp):
    shutil.rmtree(f'{folder_exp}')
    print(f"le dossier AFR_{today} à été bien supprimer et recréer\n-------------")
    print()
else:
    print(f"le dossier AFR_{today} n'existe pas\n-------------")
    print()

os.mkdir(folder_exp)

# folder_list_affiliate= f'C:/Users/J1049122/Desktop/Station Data/Master-Data/export/list_affiliate_{today}'
# os.mkdir(folder_list_affiliate)

path_data_SAP = "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/Data-SAP.xlsx"
path_data_sharepoint = "C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/all-data-sharepoint.xlsx"
path_list = f"C:/Users/J1049122/Desktop/Station Data/Master-Data/export/SAP_vs_SHAREPOINT/Total_AFR_StationData_{today}.xlsx"


def com(df_X, df_Y, col, texte = True):

    if texte:
        diff_X = np.setdiff1d(df_X[col] ,df_Y[col])
        ecart_X = df_X.loc[df_X[col].isin(diff_X)]
        
        print("Données SAP versus données Sharepoint :")
        print(f"il y'a {len(diff_X)} code SAP de différence")
        
        print()
        diff_Y = np.setdiff1d(df_Y[col], df_X[col])
        ecart_Y = df_Y.loc[df_Y[col].isin(diff_Y)]
        
        print("Données Sharepoint versus données SAP :")
        print(f"il y'a {len(diff_Y)} code SAP de différence")

        commun = df_X.loc[~df_X[col].isin(diff_X)]

        return ecart_X, ecart_Y, commun
    
    else:
        diff_X = np.setdiff1d(df_X[col] ,df_Y[col])
        ecart_X = df_X.loc[df_X[col].isin(diff_X)]

        diff_Y = np.setdiff1d(df_Y[col], df_X[col])
        ecart_Y = df_Y.loc[df_Y[col].isin(diff_Y)]
        
        commun = df_X.loc[~df_X[col].isin(diff_X)]

        return ecart_X, ecart_Y, commun        


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def comparer(Pays):

    if os.path.exists(path_list):
        os.remove(path_list)
        print("le fichier 'Affiliate_list.xlsx' à été bien supprimer et recréer\n-------------")
    else:
        print("le fichier 'Affiliate_list.xlsx' n'existe pas\n-------------")


    data_sharepoint = pd.read_excel('C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/all-data-sharepoint.xlsx')
    
    writer_list = pd.ExcelWriter(path_list, engine = 'openpyxl')

    data_sharepoint.to_excel(writer_list, sheet_name = 'Station Data Brute', index=False)
    writer_list.save()
    writer_list.close()

    print()
    All_df_sap = pd.read_excel(path_data_SAP)
    All_df_sharepoint = pd.read_excel(path_data_sharepoint)

    for i in Pays:

        element = i

        # print()

        print('-'*20)
        print(f"Pays : {element}")
        print('-'*20)

        path_ecart = f"{folder_exp}/{element}.xlsx"
        #path_list = f"{folder_list_affiliate}/list_affiliate_{str(today)}.xlsx"


        df_sap = All_df_sap[All_df_sap['Affiliate'] == element]
        df_sap.rename(columns={'SAPCODE': 'SAPCode'}, inplace=True)
        df_sap = df_sap.drop_duplicates(subset = "SAPCode", keep = 'first')
        dim_sap = df_sap.shape
        print(f"dimension données SAP pour {element} est : {dim_sap}")
        df_sap['SAPCode'] = df_sap['SAPCode'].str.strip()

        df_sharepoint = All_df_sharepoint[All_df_sharepoint['Affiliate'] == element]
        df_sharepoint = df_sharepoint.drop_duplicates()
        dim_sharepoint = df_sharepoint.shape
        print(f"dimension données sharepoint pour {element} est : {dim_sharepoint}")
        df_sharepoint['SAPCode'] = df_sharepoint['SAPCode'].str.strip()

        print()

        print("Comparaison :")
        print('-'*7)

        X, Y, df_commun_1 = com(df_sap, df_sharepoint, 'SAPCode')
        a, cost, df_commun_2 = com(df_commun_1, df_sharepoint, 'SAPCode_BM', texte=False)
        b, cost, df_commun_3 = com(df_commun_2, df_sharepoint, 'SAPCode_BM_ISACTIVESITE', texte=False)

        writer = pd.ExcelWriter(path_ecart, engine = 'openpyxl')
        df_sap.to_excel(writer, sheet_name = 'Data_SAP_Brute', index=False)
        df_sharepoint.to_excel(writer, sheet_name = 'Data_Sharepoint_Brute', index=False)
        X.to_excel(writer, sheet_name = 'ecart_SAP_vs_Sharepoint', index=False)
        Y.to_excel(writer, sheet_name = 'ecart_Sharepoint_vs_SAP', index=False)
        a.to_excel(writer, sheet_name = 'SAP_vs_Sharepoint_SAPCode_BM', index=False)
        b.to_excel(writer, sheet_name = 'SAP_vs_Sharepoint_SAPCode_BM_ISACTIVESITE', index=False)

        writer.save()
        writer.close()

        print()
        print('#'*70)
        print()


        # sh = pd.read_excel("C:/Users/J1049122/Desktop/Station Data/Master-Data/Data source/Data-SAP.xlsx")
        # sh = sh.drop_duplicates()
        # sh['SAPCode'] = sh['SAPCode'].str.strip()

        # z = sh['Affiliate'].unique()

        # # for w in z:
        # d = sh[sh['Affiliate']==w]

        ecart_sap = X.copy()

        ecart_sap = ecart_sap[["SAPCode", "Affiliate", "FINAL_SITENAME", "SITETOWN", "ISACTIVESITE", "BUSINESSMODEL", "BM_source"]]
        ecart_sap.columns = ['SAPCode', 'Affiliate', 'SAPName', 'Town', 'IsActiveSite', 'BUSINESSMODEL', 'BM_source']

        colonnes = ['Zone', 'SubZone', 'IntermediateStatus', 'Brand', 'Segment', 'ContractMode', 'ShopSegment', 'SFSActivity', 'SFSContractType', 'PartnerOrBrand', 'TargetKit', 'TargetPOSprovider', 
                    'EstimatedInstallationDate', 'InstalledSolutionOnSite', 'SolutionProvider', 'SolutionInstallationDate', 'Status', 'SolutionRelease', 'SystemOwner', 'ConfigurationStatus',
                    'IsAllPumpsConnectedToFCC', 'Reason', 'AutomaticTankGauging', 'ATGProvider', 'ATGModel', 'ATGConnected', 'ATGInstallationDate', 'TotalCardEPT connection', 'FuelCardProvider', 
                    'EPTHardware', 'EPTModel', 'EPTNumber', 'EPTConnected', 'PaymentLocation', 'HOSInstalled', 'HOSProvider', 'WSMSoftwareInstalled', 'WSMProvider', 'TELECOM', 'STABILITE TELECOM',
                    'STARTBOXStatus',  'BM_source'
                   ]

        for col in colonnes:
            ecart_sap[col] = ""
        


        all_cols_ordonner = ['SAPCode', 'Zone', 'SubZone', 'Affiliate', 'SAPName', 'Town',
            'IsActiveSite', 'IntermediateStatus', 'Brand', 'Segment',
            'BUSINESSMODEL', 'ContractMode', 'ShopSegment', 'SFSActivity',
            'SFSContractType', 'PartnerOrBrand', 'TargetKit', 'TargetPOSprovider',
            'EstimatedInstallationDate', 'InstalledSolutionOnSite',
            'SolutionProvider', 'SolutionInstallationDate', 'Status',
            'SolutionRelease', 'SystemOwner', 'ConfigurationStatus',
            'IsAllPumpsConnectedToFCC', 'Reason', 'AutomaticTankGauging',
            'ATGProvider', 'ATGModel', 'ATGConnected', 'ATGInstallationDate',
            'TotalCardEPT connection', 'FuelCardProvider', 'EPTHardware',
            'EPTModel', 'EPTNumber', 'EPTConnected', 'PaymentLocation',
            'HOSInstalled', 'HOSProvider', 'WSMSoftwareInstalled', 'WSMProvider',
            'TELECOM', 'STABILITE TELECOM', 'STARTBOXStatus', 'BM_source']

        ecart_sap1=ecart_sap.reindex(columns= all_cols_ordonner)
        ecart_sap1['data_source'] = "ecart SAP"
        ecart_sap1 = ecart_sap1[ecart_sap1['BUSINESSMODEL'] != 'CLOS']


        sh = df_sharepoint.copy()

        if a.shape[0] > 0:
            for j in range(a.shape[0]):
                for k in range(sh.shape[0]):
                    if a['SAPCode'].iloc[j] == sh['SAPCode'].iloc[k]:
                        sh['BUSINESSMODEL'].iloc[k] = a['BUSINESSMODEL'].iloc[j]
                        sh['BM_source'].iloc[k] = a['BM_source'].iloc[j]

        sh = sh[['SAPCode', 'Zone', 'SubZone', 'Affiliate', 'SAPName', 'Town','IsActiveSite', 'IntermediateStatus', 'Brand', 'Segment',
            'BUSINESSMODEL', 'ContractMode', 'ShopSegment', 'SFSActivity', 'SFSContractType', 'PartnerOrBrand', 'TargetKit', 'TargetPOSprovider',
            'EstimatedInstallationDate', 'InstalledSolutionOnSite', 'SolutionProvider', 'SolutionInstallationDate', 'Status',
            'SolutionRelease', 'SystemOwner', 'ConfigurationStatus', 'IsAllPumpsConnectedToFCC', 'Reason', 'AutomaticTankGauging',
            'ATGProvider', 'ATGModel', 'ATGConnected', 'ATGInstallationDate', 'TotalCardEPT connection', 'FuelCardProvider', 'EPTHardware',
            'EPTModel', 'EPTNumber', 'EPTConnected', 'PaymentLocation', 'HOSInstalled', 'HOSProvider', 'WSMSoftwareInstalled', 'WSMProvider',
            'TELECOM', 'STABILITE TELECOM', 'STARTBOXStatus', 'BM_source']]
        sh['data_source'] = "Station Data"

        sh_1 = sh.append(ecart_sap1, ignore_index=True)
        sh_1 = sh_1.drop_duplicates()


        book = load_workbook(path_list)
        writer_list = pd.ExcelWriter(path_list, engine = 'openpyxl')
        writer_list.book = book

        sh_1.to_excel(writer_list, sheet_name = element, index=False)
        writer_list.save()
        writer_list.close()



pays = ['Botswana', 'Ghana', 'Kenya', 'Mauritius', 'Malawi', 'Mozambique', 'Namibia',
 'Nigeria', 'Tanzania', 'Uganda', 'South Africa', 'Zambia',
 'Zimbabwe', 'Central Afr.Rep', 'Congo', 'Cameroon', 'Gabon', 'Guinea Conakry',
 'Equatorial Gui.', 'Morocco', 'Mali', 'Senegal', 'Chad', 'Togo', 'Mayotte', 'Egypt']

comparer(pays)


print()
print("--------------------")
print("Terminer avec succès")
print("--------------------")
print()

print(time.ctime(time.time() - start)[11:19])
